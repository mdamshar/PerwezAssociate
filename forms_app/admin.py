from django.contrib import admin
from django.http import HttpResponse
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from .models import QuoteRequest, Feedback

def export_quotes_to_excel(modeladmin, request, queryset):
    """Export selected quotes to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotes"
    
    # Headers
    headers = ['Name', 'Phone', 'Address', 'Topic', 'Message', 'Status', 'Created At', 'Updated At']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Add data
    for quote in queryset:
        ws.append([
            quote.name,
            quote.phone,
            quote.address,
            quote.get_topic_display(),
            quote.message,
            quote.get_status_display(),
            quote.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            quote.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="quotes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response

export_quotes_to_excel.short_description = "📥 Download selected quotes as Excel"


def export_accepted_quotes_to_excel(modeladmin, request, queryset):
    """Export accepted quotes to Excel"""
    # Filter only accepted quotes
    accepted_quotes = queryset.filter(status='accepted')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Accepted Quotes"
    
    # Headers
    headers = ['Name', 'Phone', 'Address', 'Topic', 'Message', 'Accepted Date', 'Notes']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Add data
    for quote in accepted_quotes:
        ws.append([
            quote.name,
            quote.phone,
            quote.address,
            quote.get_topic_display(),
            quote.message,
            quote.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
            "Accepted Quote",
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="accepted_quotes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response

export_accepted_quotes_to_excel.short_description = "✅ Download accepted quotes as Excel"


def mark_as_accepted(modeladmin, request, queryset):
    """Mark quotes as accepted"""
    updated = queryset.update(status='accepted')
    modeladmin.message_user(request, f'✅ {updated} quote(s) marked as accepted.')


mark_as_accepted.short_description = "✅ Mark as Accepted"


def mark_as_deleted(modeladmin, request, queryset):
    """Mark quotes as deleted"""
    updated = queryset.update(status='deleted')
    modeladmin.message_user(request, f'🗑️ {updated} quote(s) marked as deleted.')

mark_as_deleted.short_description = "🗑️ Mark as Deleted"


@admin.register(QuoteRequest)
class QuoteRequestAdmin(admin.ModelAdmin):
    list_display = ('name', 'phone', 'topic', 'status_badge', 'created_at')
    list_filter = ('status', 'topic', 'created_at')
    search_fields = ('name', 'phone', 'address')
    readonly_fields = ('created_at', 'updated_at')
    actions = [mark_as_accepted, mark_as_deleted, export_quotes_to_excel, export_accepted_quotes_to_excel]
    
    fieldsets = (
        ('📋 Contact Information', {
            'fields': ('name', 'phone', 'address')
        }),
        ('📝 Request Details', {
            'fields': ('topic', 'message', 'status')
        }),
        ('⏰ Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def status_badge(self, obj):
        colors = {
            'pending': '#FFC107',
            'accepted': '#28A745',
            'deleted': '#DC3545'
        }
        text_colors = {
            'pending': '#333333',
            'accepted': 'white',
            'deleted': 'white'
        }
        icons = {
            'pending': '⏳',
            'accepted': '✅',
            'deleted': '🗑️'
        }
        color = colors.get(obj.status, '#999999')
        text_color = text_colors.get(obj.status, 'white')
        icon = icons.get(obj.status, '?')
        return f"<span style='background: {color}; color: {text_color}; padding: 5px 10px; border-radius: 5px;'>{icon} {obj.get_status_display()}</span>"
    
    status_badge.short_description = 'Status'
    status_badge.allow_tags = True
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.exclude(status='deleted')


@admin.register(Feedback)
class FeedbackAdmin(admin.ModelAdmin):
    list_display = ('name', 'mobile', 'preview', 'created_at')
    list_filter = ('created_at',)
    search_fields = ('name', 'mobile', 'message')
    readonly_fields = ('created_at', 'message_display')
    
    fieldsets = (
        ('👤 Customer Info', {
            'fields': ('name', 'mobile')
        }),
        ('💬 Feedback', {
            'fields': ('message_display',)
        }),
        ('⏰ Timestamp', {
            'fields': ('created_at',),
            'classes': ('collapse',)
        }),
    )
    
    def preview(self, obj):
        return obj.message[:50] + '...' if len(obj.message) > 50 else obj.message
    preview.short_description = 'Message Preview'
    
    def message_display(self, obj):
        return f'<p style="white-space: pre-wrap;">{obj.message}</p>'
    message_display.short_description = 'Message'
    message_display.allow_tags = True
    
    def has_add_permission(self, request):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return True
